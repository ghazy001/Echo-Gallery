from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Artwork, ArtworkFeedback
from .forms import ArtworkForm, ArtworkFeedbackForm
from django.db.models import Q
from django.core.paginator import Paginator
from django.utils.dateparse import parse_date
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from django.http import HttpResponse
from django.db.models import Count, Avg, Q, F
from django.db.models.functions import TruncDate
from django.utils import timezone
import json
from collections import defaultdict



def staff_only(request):
    return request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser)

# ------------ ARTWORK CRUD ------------

@login_required
def artwork_list_view(request):
    if not staff_only(request):
        return redirect('profile')

    # --- GET filters ---
    search_q      = request.GET.get('q', '').strip()
    visible_q     = request.GET.get('visible', '').strip()  # "yes" / "no" / ""
    year_from     = request.GET.get('year_from', '').strip()
    year_to       = request.GET.get('year_to', '').strip()
    created_from  = request.GET.get('created_from', '').strip()
    created_to    = request.GET.get('created_to', '').strip()

    # --- base queryset ---
    qs = Artwork.objects.all()

    # text search by title, artist, description, year
    if search_q:
        qs = qs.filter(
            Q(title__icontains=search_q) |
            Q(artist__icontains=search_q) |
            Q(description__icontains=search_q) |
            Q(year__icontains=search_q)
        )

    # visible filter
    if visible_q == "yes":
        qs = qs.filter(is_visible=True)
    elif visible_q == "no":
        qs = qs.filter(is_visible=False)

    # year range is tricky, because `year` is CharField (not numeric)
    # We'll do a simple contains-ish filter:
    # If admin types "1800", show anything whose year contains "1800".
    if year_from:
        qs = qs.filter(year__icontains=year_from)
    if year_to and year_to != year_from:
        qs = qs.filter(year__icontains=year_to)

    # created_at date range
    # parse_date => returns datetime.date
    if created_from:
        dfrom = parse_date(created_from)
        if dfrom:
            qs = qs.filter(created_at__date__gte=dfrom)

    if created_to:
        dto = parse_date(created_to)
        if dto:
            qs = qs.filter(created_at__date__lte=dto)

    # default ordering is handled by Meta.ordering = ['-created_at'],
    # but we can be explicit too:
    qs = qs.order_by('-created_at')

    # paginate
    p = Paginator(qs, 20)
    page_number = request.GET.get('page')
    page_obj = p.get_page(page_number)

    return render(request, 'dashboard/artwork/artwork_list.html', {
        'artworks': page_obj,
        'page_obj': page_obj,

        # sticky filters for the template
        'filter_q': search_q,
        'filter_visible': visible_q,
        'filter_year_from': year_from,
        'filter_year_to': year_to,
        'filter_created_from': created_from,
        'filter_created_to': created_to,
    })


@login_required
def artwork_create_view(request):
    if not staff_only(request):
        return redirect('profile')

    if request.method == "POST":
        form = ArtworkForm(request.POST, request.FILES)   # <-- ADD request.FILES
        if form.is_valid():
            form.save()
            return redirect('dashboard_artwork_list')
    else:
        form = ArtworkForm()

    return render(request, 'dashboard/artwork/artwork_form.html', {
        'form': form,
        'form_title': 'Add Artwork',
        'submit_label': 'Create',
    })


@login_required
def artwork_edit_view(request, artwork_id):
    if not staff_only(request):
        return redirect('profile')

    artwork = get_object_or_404(Artwork, id=artwork_id)

    if request.method == "POST":
        form = ArtworkForm(request.POST, request.FILES, instance=artwork)  # <-- files + instance
        if form.is_valid():
            form.save()
            return redirect('dashboard_artwork_list')
    else:
        form = ArtworkForm(instance=artwork)

    return render(request, 'dashboard/artwork/artwork_form.html', {
        'form': form,
        'form_title': f'Edit Artwork: {artwork.title}',
        'submit_label': 'Save',
    })


@login_required
def artwork_delete_view(request, artwork_id):
    if not staff_only(request):
        return redirect('profile')

    artwork = get_object_or_404(Artwork, id=artwork_id)

    if request.method == "POST":
        artwork.delete()
        return redirect('dashboard_artwork_list')

    return render(request, 'dashboard/confirm_delete.html', {
        'object_name': artwork.title,
        'cancel_url_name': 'dashboard_artwork_list',
    })


# ------------ FEEDBACK MODERATION CRUD ------------

@login_required
def feedback_list_view(request):
    if not staff_only(request):
        return redirect('profile')

    # newest first, so admin sees recent comments
    feedbacks = (
        ArtworkFeedback.objects
        .select_related('artwork', 'user')
        .order_by('-created_at')
    )

    return render(request, 'dashboard/feedback/feedback_list.html', {
        'feedbacks': feedbacks,
    })


@login_required
def feedback_edit_view(request, feedback_id):
    if not staff_only(request):
        return redirect('profile')

    fb = get_object_or_404(ArtworkFeedback, id=feedback_id)

    if request.method == "POST":
        form = ArtworkFeedbackForm(request.POST, instance=fb)
        if form.is_valid():
            form.save()
            return redirect('dashboard_feedback_list')
    else:
        form = ArtworkFeedbackForm(instance=fb)

    return render(request, 'dashboard/feedback/feedback_form.html', {
        'form': form,
        'form_title': f'Edit Feedback on {fb.artwork.title}',
        'submit_label': 'Save',
    })


@login_required
def feedback_delete_view(request, feedback_id):
    if not staff_only(request):
        return redirect('profile')

    fb = get_object_or_404(ArtworkFeedback, id=feedback_id)

    if request.method == "POST":
        fb.delete()
        return redirect('dashboard_feedback_list')

    return render(request, 'dashboard/confirm_delete.html', {
        'object_name': f'Feedback on {fb.artwork.title}',
        'cancel_url_name': 'dashboard_feedback_list',
    })



# ------------ EXPORT TO EXCEL ------------
@login_required
def feedback_export_excel_view(request):
    if not staff_only(request):
        return redirect('profile')

    # Pull same data as feedback_list_view (newest first)
    feedbacks = (
        ArtworkFeedback.objects
        .select_related('artwork', 'user')
        .order_by('-created_at')
    )

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feedback"

    # Define headers
    headers = [
        "Artwork",
        "User / Name",
        "Rating",
        "Approved?",
        "Comment",
        "Created At",
    ]

    # Write header row with bold + light fill
    header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = 28

    # Freeze header row for scrolling
    ws.freeze_panes = "A2"

    # Write data rows
    row_idx = 2
    for fb in feedbacks:
        # Figure out who left the feedback
        if fb.user:
            user_display = fb.user.username
        elif fb.name:
            user_display = fb.name
        else:
            user_display = "Anonymous"

        # Approved styling cell
        approved_text = "YES" if fb.is_approved else "Pending"
        approved_cell_fill = PatternFill(
            start_color=("C6EFCE" if fb.is_approved else "F2F2F2"),
            end_color=("C6EFCE" if fb.is_approved else "F2F2F2"),
            fill_type="solid"
        )

        # Created timestamp (string, so Excel shows it nicely)
        created_str = fb.created_at.strftime("%Y-%m-%d %H:%M")

        # Comment body (full, not truncated)
        comment_text = fb.comment or ""

        ws.cell(row=row_idx, column=1, value=fb.artwork.title if fb.artwork else "")
        ws.cell(row=row_idx, column=2, value=user_display)
        ws.cell(row=row_idx, column=3, value=f"{fb.rating}/5" if fb.rating is not None else "")
        ws.cell(row=row_idx, column=4, value=approved_text)
        ws.cell(row=row_idx, column=5, value=comment_text)
        ws.cell(row=row_idx, column=6, value=created_str)

        # apply fill to the Approved? cell we just wrote (col 4)
        ws.cell(row=row_idx, column=4).fill = approved_cell_fill

        row_idx += 1

    # Response as download
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=feedback.xlsx'

    wb.save(response)
    return response



# ------------ STATS ------------

@login_required
def artwork_stats_view(request):
    if not staff_only(request):
        return redirect('profile')

    now = timezone.now()

    # ---------- BASE QS ----------
    artworks_qs = Artwork.objects.all()
    feedback_qs = ArtworkFeedback.objects.select_related('artwork').all()

    total_artworks = artworks_qs.count()
    visible_artworks = artworks_qs.filter(is_visible=True).count()
    hidden_artworks = artworks_qs.filter(is_visible=False).count()

    # recently added: last 30 days
    last_30_days = now - timezone.timedelta(days=30)
    new_last_30_days = artworks_qs.filter(created_at__gte=last_30_days).count()

    # feedback stats
    total_feedback = feedback_qs.count()
    approved_feedback = feedback_qs.filter(is_approved=True).count()

    # average rating (only approved, since that's what you'd actually show publicly)
    avg_rating_val = (
        feedback_qs
        .filter(is_approved=True, rating__isnull=False)
        .aggregate(avg_rating=Avg('rating'))
        ['avg_rating']
    )
    avg_rating = round(avg_rating_val, 2) if avg_rating_val is not None else 0

    # top artworks by avg rating (only consider approved feedback, and only those with >=1 approved feedback)
    top_rated_artworks = (
        Artwork.objects
        .annotate(
            approved_feedback_count=Count('feedbacks', filter=Q(feedbacks__is_approved=True)),
            avg_rating=Avg('feedbacks__rating', filter=Q(feedbacks__is_approved=True, feedbacks__rating__isnull=False)),
        )
        .filter(approved_feedback_count__gt=0)
        .order_by('-avg_rating', '-approved_feedback_count', 'title')[:10]
    )

    # artworks most commented overall
    most_commented_artworks = (
        Artwork.objects
        .annotate(fb_count=Count('feedbacks'))
        .filter(fb_count__gt=0)
        .order_by('-fb_count', 'title')[:10]
    )

    # --------- CHART DATA 1: Artworks per "year" ---------
    # year is a CharField (e.g. "1889", "c. 5th century BC", etc.)
    # We'll count by exact string. We'll take top ~10 most common "years".
    artworks_per_year_raw = (
        artworks_qs
        .values('year')
        .annotate(cnt=Count('id'))
        .order_by('-cnt', 'year')
    )

    # Keep only non-empty and take first 10 buckets
    artworks_per_year = [row for row in artworks_per_year_raw if row['year'].strip()][:10]

    chart_year_labels = [row['year'] for row in artworks_per_year]
    chart_year_counts = [row['cnt'] for row in artworks_per_year]

    # --------- CHART DATA 2: Feedback per day (last 14 days) ---------
    last_14_days = now - timezone.timedelta(days=14)
    fb_last_14 = (
        feedback_qs
        .filter(created_at__gte=last_14_days)
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(cnt=Count('id'))
        .order_by('day')
    )

    # build complete series (even days with 0)
    fb_counts_by_day = {row['day'].isoformat(): row['cnt'] for row in fb_last_14}
    timeline_labels = []
    timeline_counts = []
    for i in range(15):  # inclusive today
        d = (last_14_days + timezone.timedelta(days=i)).date()
        key = d.isoformat()
        timeline_labels.append(key)
        timeline_counts.append(fb_counts_by_day.get(key, 0))

    # We'll serialize chart data into JSON for safe embedding into the template.
    chart_context = {
        'chart_year_labels': json.dumps(chart_year_labels),
        'chart_year_counts': json.dumps(chart_year_counts),
        'timeline_labels': json.dumps(timeline_labels),
        'timeline_counts': json.dumps(timeline_counts),
    }

    context = {
        'now': now,

        # KPIs
        'total_artworks': total_artworks,
        'visible_artworks': visible_artworks,
        'hidden_artworks': hidden_artworks,
        'new_last_30_days': new_last_30_days,
        'total_feedback': total_feedback,
        'approved_feedback': approved_feedback,
        'avg_rating': avg_rating,

        # tables
        'top_rated_artworks': top_rated_artworks,
        'most_commented_artworks': most_commented_artworks,

        # charts (as JSON strings ready for JS)
        **chart_context,
    }

    return render(request, 'dashboard/artwork/artwork_stats.html', context)