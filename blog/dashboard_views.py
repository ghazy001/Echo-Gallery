from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.utils.dateparse import parse_date
from django.core.paginator import Paginator

from .models import Category, Article
from .forms import CategoryForm, ArticleForm
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count
from django.db.models.functions import TruncMonth
import calendar


def staff_only(request):
    return request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser)

# ------------- CATEGORY CRUD -------------

@login_required
def category_list_view(request):
    if not staff_only(request):
        return redirect('profile')

    categories = Category.objects.all().order_by('name')
    return render(request, 'dashboard/category/category_list.html', {
        'categories': categories,
    })


@login_required
def category_create_view(request):
    if not staff_only(request):
        return redirect('profile')

    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('dashboard_category_list')
    else:
        form = CategoryForm()

    return render(request, 'dashboard/category/category_form.html', {
        'form': form,
        'form_title': 'Add Category',
        'submit_label': 'Create',
    })


@login_required
def category_edit_view(request, category_id):
    if not staff_only(request):
        return redirect('profile')

    category = get_object_or_404(Category, id=category_id)

    if request.method == "POST":
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('dashboard_category_list')
    else:
        form = CategoryForm(instance=category)

    return render(request, 'dashboard/category/category_form.html', {
        'form': form,
        'form_title': f'Edit Category: {category.name}',
        'submit_label': 'Save',
    })


@login_required
def category_delete_view(request, category_id):
    if not staff_only(request):
        return redirect('profile')

    category = get_object_or_404(Category, id=category_id)

    if request.method == "POST":
        # Will raise ProtectedError if there are still articles using it
        category.delete()
        return redirect('dashboard_category_list')

    return render(request, 'dashboard/confirm_delete.html', {
        'object_name': category.name,
        'cancel_url_name': 'dashboard_category_list',
    })


# ------------- ARTICLE CRUD -------------

@login_required
def article_list_view(request):
    if not staff_only(request):
        return redirect('profile')

    # --- read filters from querystring ---
    search_q      = request.GET.get('q', '').strip()
    category_id   = request.GET.get('category', '').strip()
    author_id     = request.GET.get('author', '').strip()
    published     = request.GET.get('published', '').strip()  # "yes", "no", or ""
    date_from_raw = request.GET.get('date_from', '').strip()
    date_to_raw   = request.GET.get('date_to', '').strip()

    # --- base queryset ---
    qs = (
        Article.objects
        .select_related('category', 'author')
        .all()
    )

    # --- keyword search (title / summary / body) ---
    if search_q:
        qs = qs.filter(
            Q(title__icontains=search_q) |
            Q(summary__icontains=search_q) |
            Q(body__icontains=search_q)
        )

    # --- category filter ---
    if category_id:
        qs = qs.filter(category_id=category_id)

    # --- author filter ---
    if author_id:
        qs = qs.filter(author_id=author_id)

    # --- published filter ---
    if published == "yes":
        qs = qs.filter(is_published=True)
    elif published == "no":
        qs = qs.filter(is_published=False)

    # --- date range on published_at ---
    if date_from_raw:
        dfrom = parse_date(date_from_raw)
        if dfrom:
            qs = qs.filter(published_at__date__gte=dfrom)

    if date_to_raw:
        dto = parse_date(date_to_raw)
        if dto:
            qs = qs.filter(published_at__date__lte=dto)

    # newest first (also enforced by Meta.ordering)
    qs = qs.order_by('-published_at')

    # --- dropdown data sources ---
    categories = Category.objects.all().order_by('name')

    authors = (
        Article.objects
        .select_related('author')
        .values('author_id', 'author__username')
        .distinct()
        .order_by('author__username')
    )

    # --- pagination ---
    p = Paginator(qs, 20)  # 20 rows per page
    page_number = request.GET.get('page')
    page_obj = p.get_page(page_number)

    # --- render ---
    return render(request, 'dashboard/articles/article_list.html', {
        'articles': page_obj,     # this is iterable in template
        'page_obj': page_obj,     # for pagination controls

        'categories': categories,
        'authors': authors,

        # stick filters in the form
        'filter_q': search_q,
        'filter_category': category_id,
        'filter_author': author_id,
        'filter_published': published,
        'filter_date_from': date_from_raw,
        'filter_date_to': date_to_raw,
    })


@login_required
def article_create_view(request):
    if not staff_only(request):
        return redirect('profile')

    if request.method == "POST":
        form = ArticleForm(request.POST, request.FILES)  # 👈 add request.FILES
        if form.is_valid():
            article = form.save(commit=False)
            if article.author_id is None:  # if not set by form
                article.author = request.user
            article.save()
            return redirect('dashboard_article_list')
    else:
        form = ArticleForm()

    return render(request, 'dashboard/articles/article_form.html', {
        'form': form,
        'form_title': 'Add Article',
        'submit_label': 'Create',
    })


@login_required
def article_edit_view(request, article_id):
    if not staff_only(request):
        return redirect('profile')

    article = get_object_or_404(Article, id=article_id)

    if request.method == "POST":
        form = ArticleForm(request.POST, request.FILES, instance=article)  #
        if form.is_valid():
            form.save()
            return redirect('dashboard_article_list')
    else:
        form = ArticleForm(instance=article)

    return render(request, 'dashboard/articles/article_form.html', {
        'form': form,
        'form_title': f'Edit Article: {article.title}',
        'submit_label': 'Save',
    })


@login_required
def article_delete_view(request, article_id):
    if not staff_only(request):
        return redirect('profile')

    article = get_object_or_404(Article, id=article_id)

    if request.method == "POST":
        article.delete()
        return redirect('dashboard_article_list')

    return render(request, 'dashboard/confirm_delete.html', {
        'object_name': article.title,
        'cancel_url_name': 'dashboard_article_list',
    })


# -------- EXPORT TO EXCEL --------
@login_required
def category_export_excel_view(request):
    if not staff_only(request):
        return redirect('profile')

    # 1. Query data
    categories = Category.objects.all().order_by('name')

    # 2. Create workbook + sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categories"

    # 3. Header row
    headers = ["Name", "Slug", "Articles"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        # basic bold/style
        cell.font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # 4. Data rows
    row_idx = 2
    for c in categories:
        ws.cell(row=row_idx, column=1, value=c.name)
        ws.cell(row=row_idx, column=2, value=c.slug)
        ws.cell(row=row_idx, column=3, value=c.articles.count())
        row_idx += 1

    # 5. Build HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=categories.xlsx'

    wb.save(response)
    return response



#-------- STATS --------
@login_required
def article_stats_view(request):
    if not staff_only(request):
        return redirect('profile')

    now = timezone.now()

    # base queryset
    qs = (
        Article.objects
        .select_related('category', 'author')
        .all()
    )

    total_articles = qs.count()
    published_articles = qs.filter(is_published=True).count()
    draft_articles = qs.filter(is_published=False).count()

    # recently published metrics
    # "recent" here = published_at within last 30 days
    last_30_days = now - timezone.timedelta(days=30)
    articles_last_30_days = qs.filter(published_at__gte=last_30_days).count()

    # % published
    published_pct = 0
    if total_articles > 0:
        published_pct = round((published_articles / total_articles) * 100)

    # articles per category
    articles_by_category = (
        qs.values('category__id', 'category__name')
          .annotate(article_count=Count('id'))
          .order_by('-article_count', 'category__name')
    )
    # shape: [{'category__id': 1, 'category__name': 'News', 'article_count': 12}, ...]

    # articles per author
    # we also protect against null author (it’s allowed in your model)
    articles_by_author = (
        qs.values('author_id', 'author__username')
          .annotate(article_count=Count('id'))
          .order_by('-article_count', 'author__username')
    )
    # shape: [{'author_id': 5, 'author__username': 'alice', 'article_count': 7}, ...]

    # articles per month (based on published_at)
    raw_month_counts = (
        qs.annotate(month=TruncMonth('published_at'))
          .values('month')
          .annotate(article_count=Count('id'))
          .order_by('month')
    )

    articles_by_month = []
    for row in raw_month_counts:
        month_start = row['month'].date().replace(day=1)
        last_day_num = calendar.monthrange(month_start.year, month_start.month)[1]
        month_end = month_start.replace(day=last_day_num)

        articles_by_month.append({
            'month': row['month'],
            'article_count': row['article_count'],
            'range_start': month_start.strftime("%Y-%m-%d"),
            'range_end': month_end.strftime("%Y-%m-%d"),
        })

    context = {
        'now': now,

        'total_articles': total_articles,
        'published_articles': published_articles,
        'draft_articles': draft_articles,
        'articles_last_30_days': articles_last_30_days,
        'published_pct': published_pct,

        'articles_by_category': articles_by_category,
        'articles_by_author': articles_by_author,
        'articles_by_month': articles_by_month,
    }

    return render(request, 'dashboard/articles/article_stats.html', context)