# events/dashboard_views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from .models import Place, Event
from .forms import PlaceForm, EventForm
from django.utils.dateparse import parse_date
from django.core.paginator import Paginator
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.styles import Font
from django.utils import timezone
from django.db.models import Count
from django.db.models.functions import TruncMonth

def staff_only(request):
    return request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser)

# ---------- PLACES ----------

@login_required
def place_list_view(request):
    if not staff_only(request):
        return redirect('profile')

    places = Place.objects.all().order_by('name')
    return render(request, 'dashboard/places/place_list.html', {
        'places': places,
    })


@login_required
def place_create_view(request):
    if not staff_only(request):
        return redirect('profile')

    if request.method == "POST":
        form = PlaceForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('dashboard_place_list')
    else:
        form = PlaceForm()

    return render(request, 'dashboard/places/place_form.html', {
        'form': form,
        'form_title': 'Add Place',
        'submit_label': 'Create',
    })


@login_required
def place_edit_view(request, place_id):
    if not staff_only(request):
        return redirect('profile')

    place = get_object_or_404(Place, id=place_id)

    if request.method == "POST":
        form = PlaceForm(request.POST, instance=place)
        if form.is_valid():
            form.save()
            return redirect('dashboard_place_list')
    else:
        form = PlaceForm(instance=place)

    return render(request, 'dashboard/places/place_form.html', {
        'form': form,
        'form_title': f'Edit Place: {place.name}',
        'submit_label': 'Save',
    })


@login_required
def place_delete_view(request, place_id):
    if not staff_only(request):
        return redirect('profile')

    place = get_object_or_404(Place, id=place_id)

    if request.method == "POST":
        # because of on_delete=PROTECT in Event.place,
        # this will raise error if there are still events using this place
        place.delete()
        return redirect('dashboard_place_list')

    return render(request, 'dashboard/confirm_delete.html', {
        'object_name': place.name,
        'cancel_url_name': 'dashboard_place_list',
    })


# ---------- EVENTS ----------

@login_required
def event_list_view(request):
    if not staff_only(request):
        return redirect('profile')

    # --- read filters from querystring ---
    search_q      = request.GET.get('q', '').strip()
    place_id      = request.GET.get('place', '').strip()
    published     = request.GET.get('published', '').strip()  # "yes", "no", or ""
    date_from_raw = request.GET.get('date_from', '').strip()  # filter start_date >= this date
    date_to_raw   = request.GET.get('date_to', '').strip()    # filter end_date   <= this date

    # --- base queryset ---
    qs = (
        Event.objects
        .select_related('place')
        .all()
    )

    # --- text search ---
    if search_q:
        qs = qs.filter(
            Q(title__icontains=search_q) |
            Q(description__icontains=search_q)
        )

    # --- place filter ---
    if place_id:
        qs = qs.filter(place_id=place_id)

    # --- published filter ---
    if published == "yes":
        qs = qs.filter(is_published=True)
    elif published == "no":
        qs = qs.filter(is_published=False)

    # --- date range filters ---
    # We treat date_from as "events starting on/after this day"
    # and date_to as "events ending on/before this day"
    if date_from_raw:
        dfrom = parse_date(date_from_raw)  # returns datetime.date
        if dfrom:
            qs = qs.filter(start_date__date__gte=dfrom)

    if date_to_raw:
        dto = parse_date(date_to_raw)
        if dto:
            qs = qs.filter(end_date__date__lte=dto)

    # --- sort upcoming first (soonest first) OR newest first?
    # For dashboard, it's often nicer to see most recent / upcoming first.
    # We'll sort by -start_date like you had.
    qs = qs.order_by('-start_date')

    # --- dropdown data sources ---
    places = Place.objects.all().order_by('name')

    # --- pagination ---
    p = Paginator(qs, 20)  # 20 rows per page
    page_number = request.GET.get('page')
    page_obj = p.get_page(page_number)

    # --- render ---
    return render(request, 'dashboard/events/event_list.html', {
        'events': page_obj,
        'page_obj': page_obj,
        'places': places,

        # pass current filters to keep form sticky + persist in pagination links
        'filter_q': search_q,
        'filter_place': place_id,
        'filter_published': published,
        'filter_date_from': date_from_raw,
        'filter_date_to': date_to_raw,
    })


@login_required
def event_create_view(request):
    if not staff_only(request):
        return redirect('profile')

    if request.method == "POST":
        form = EventForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('dashboard_event_list')
    else:
        form = EventForm()

    return render(request, 'dashboard/events/event_form.html', {
        'form': form,
        'form_title': 'Add Event',
        'submit_label': 'Create',
    })


@login_required
def event_edit_view(request, event_id):
    if not staff_only(request):
        return redirect('profile')

    event = get_object_or_404(Event, id=event_id)

    if request.method == "POST":
        form = EventForm(request.POST, instance=event)
        if form.is_valid():
            form.save()
            return redirect('dashboard_event_list')
    else:
        form = EventForm(instance=event)

    return render(request, 'dashboard/events/event_form.html', {
        'form': form,
        'form_title': f'Edit Event: {event.title}',
        'submit_label': 'Save',
    })


@login_required
def event_delete_view(request, event_id):
    if not staff_only(request):
        return redirect('profile')

    event = get_object_or_404(Event, id=event_id)

    if request.method == "POST":
        event.delete()
        return redirect('dashboard_event_list')

    return render(request, 'dashboard/confirm_delete.html', {
        'object_name': event.title,
        'cancel_url_name': 'dashboard_event_list',
    })


# -------- EXPORT TO EXCEL --------

@login_required
def place_export_excel_view(request):
    if not staff_only(request):
        return redirect('profile')

    places = Place.objects.all().order_by('name')

    # workbook + sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Places"

    # header row
    headers = ["Name", "Address", "City", "Capacity", "Events count"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # freeze header row for convenience
    ws.freeze_panes = "A2"

    # data rows
    row_idx = 2
    for p in places:
        ws.cell(row=row_idx, column=1, value=p.name)
        ws.cell(row=row_idx, column=2, value=p.address)
        ws.cell(row=row_idx, column=3, value=p.city)
        ws.cell(row=row_idx, column=4, value=p.capacity)
        # if you have related_name='events' on Event.place, this will work:
        ws.cell(row=row_idx, column=5, value=p.event_set.count() if hasattr(p, "event_set") else None)
        row_idx += 1

    # build response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=places.xlsx'

    wb.save(response)
    return response



# ------ Statistics ------
@login_required
def event_stats_view(request):
    if not staff_only(request):
        return redirect('profile')

    now = timezone.now()

    # base queryset (we'll reuse it a lot)
    qs = Event.objects.select_related('place').all()

    total_events = qs.count()
    published_events = qs.filter(is_published=True).count()
    unpublished_events = qs.filter(is_published=False).count()

    upcoming_events = qs.filter(start_date__gte=now).count()
    past_events = qs.filter(end_date__lt=now).count()

    # events per place, sorted by most events first
    events_by_place = (
        qs.values('place__id', 'place__name')
          .annotate(event_count=Count('id'))
          .order_by('-event_count', 'place__name')
    )
    # shape: [{'place__id': 3, 'place__name': 'Town Hall', 'event_count': 12}, ...]

    # events per month (based on start_date)
    # we'll take last 12 months including this month
    events_by_month = (
        qs.annotate(month=TruncMonth('start_date'))
          .values('month')
          .annotate(event_count=Count('id'))
          .order_by('month')
    )
    # shape: [{'month': datetime(2025, 1, 1, 0, 0, tzinfo=...), 'event_count': 7}, ...]

    context = {
        'total_events': total_events,
        'published_events': published_events,
        'unpublished_events': unpublished_events,
        'upcoming_events': upcoming_events,
        'past_events': past_events,
        'events_by_place': events_by_place,
        'events_by_month': events_by_month,
        'now': now,
    }

    return render(request, 'dashboard/events/event_stats.html', context)