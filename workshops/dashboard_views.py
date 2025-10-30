from django.shortcuts import render, redirect, get_object_or_404
from .forms import WorkshopForm, MaterialForm
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db.models import Q
from django.core.paginator import Paginator
from django.utils.dateparse import parse_date
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.http import HttpResponse
from django.db.models import Count, Avg
from django.db.models.functions import TruncMonth
from django.utils import timezone
import calendar


from .models import Workshop, Material
from events.models import Place  # you already import this in models, reuse it

def staff_only(request):
    return request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser)

# ----- MATERIAL CRUD -----

@login_required
def material_list_view(request):
    if not staff_only(request):
        return redirect('profile')

    materials = Material.objects.all().order_by('name')
    return render(request, 'dashboard/material/material_list.html', {
        'materials': materials,
    })


@login_required
def material_create_view(request):
    if not staff_only(request):
        return redirect('profile')

    if request.method == "POST":
        form = MaterialForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('dashboard_material_list')
    else:
        form = MaterialForm()

    return render(request, 'dashboard/material/material_form.html', {
        'form': form,
        'form_title': 'Add Material',
        'submit_label': 'Create',
    })


@login_required
def material_edit_view(request, material_id):
    if not staff_only(request):
        return redirect('profile')

    material = get_object_or_404(Material, id=material_id)

    if request.method == "POST":
        form = MaterialForm(request.POST, instance=material)
        if form.is_valid():
            form.save()
            return redirect('dashboard_material_list')
    else:
        form = MaterialForm(instance=material)

    return render(request, 'dashboard/material/material_form.html', {
        'form': form,
        'form_title': f'Edit Material: {material.name}',
        'submit_label': 'Save',
    })


@login_required
def material_delete_view(request, material_id):
    if not staff_only(request):
        return redirect('profile')

    material = get_object_or_404(Material, id=material_id)

    if request.method == "POST":
        material.delete()
        return redirect('dashboard_material_list')

    return render(request, 'dashboard/confirm_delete.html', {
        'object_name': material.name,
        'cancel_url_name': 'dashboard_material_list',
    })


# ----- WORKSHOP CRUD -----




@login_required
def workshop_list_view(request):
    if not staff_only(request):
        return redirect('profile')

    # --- read filters from GET ---
    search_q        = request.GET.get('q', '').strip()
    place_id        = request.GET.get('place', '').strip()
    active_filter   = request.GET.get('active', '').strip()  # "yes", "no", or ""
    cap_min_raw     = request.GET.get('cap_min', '').strip()
    cap_max_raw     = request.GET.get('cap_max', '').strip()
    date_from_raw   = request.GET.get('date_from', '').strip()  # start_time >= this date
    date_to_raw     = request.GET.get('date_to', '').strip()    # end_time   <= this date

    # --- base queryset ---
    qs = (
        Workshop.objects
        .select_related('place')
        .prefetch_related('materials')
        .all()
    )

    # --- text search: title / instructor / description ---
    if search_q:
        qs = qs.filter(
            Q(title__icontains=search_q) |
            Q(instructor__icontains=search_q) |
            Q(description__icontains=search_q)
        )

    # --- place filter ---
    if place_id:
        qs = qs.filter(place_id=place_id)

    # --- active filter ---
    # is_active acts like "visible / hidden"
    if active_filter == "yes":
        qs = qs.filter(is_active=True)
    elif active_filter == "no":
        qs = qs.filter(is_active=False)

    # --- capacity min/max ---
    # NOTE: we only apply if the value is a valid integer
    if cap_min_raw.isdigit():
        qs = qs.filter(capacity__gte=int(cap_min_raw))
    if cap_max_raw.isdigit():
        qs = qs.filter(capacity__lte=int(cap_max_raw))

    # --- date range filters (by day portion of start_time / end_time) ---
    if date_from_raw:
        dfrom = parse_date(date_from_raw)
        if dfrom:
            qs = qs.filter(start_time__date__gte=dfrom)

    if date_to_raw:
        dto = parse_date(date_to_raw)
        if dto:
            qs = qs.filter(end_time__date__lte=dto)

    # --- order newest first (like Events) ---
    qs = qs.order_by('-start_time')

    # --- dropdown data: places list for the filter select ---
    places = Place.objects.all().order_by('name')

    # --- pagination ---
    p = Paginator(qs, 20)  # 20 rows per page
    page_number = request.GET.get('page')
    page_obj = p.get_page(page_number)

    # --- render ---
    return render(request, 'dashboard/workshops/workshop_list.html', {
        'workshops': page_obj,
        'page_obj': page_obj,
        'places': places,

        # sticky filters for template
        'filter_q': search_q,
        'filter_place': place_id,
        'filter_active': active_filter,
        'filter_cap_min': cap_min_raw,
        'filter_cap_max': cap_max_raw,
        'filter_date_from': date_from_raw,
        'filter_date_to': date_to_raw,
    })



@login_required
def workshop_create_view(request):
    if not staff_only(request):
        return redirect('profile')

    if request.method == "POST":
        form = WorkshopForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('dashboard_workshop_list')
    else:
        form = WorkshopForm()

    return render(request, 'dashboard/workshops/workshop_form.html', {
        'form': form,
        'form_title': 'Add Workshop',
        'submit_label': 'Create',
    })


@login_required
def workshop_edit_view(request, workshop_id):
    if not staff_only(request):
        return redirect('profile')

    workshop = get_object_or_404(Workshop, id=workshop_id)

    if request.method == "POST":
        form = WorkshopForm(request.POST, instance=workshop)
        if form.is_valid():
            form.save()
            return redirect('dashboard_workshop_list')
    else:
        form = WorkshopForm(instance=workshop)

    return render(request, 'dashboard/workshops/workshop_form.html', {
        'form': form,
        'form_title': f'Edit Workshop: {workshop.title}',
        'submit_label': 'Save',
    })


@login_required
def workshop_delete_view(request, workshop_id):
    if not staff_only(request):
        return redirect('profile')

    workshop = get_object_or_404(Workshop, id=workshop_id)

    if request.method == "POST":
        workshop.delete()
        return redirect('dashboard_workshop_list')

    return render(request, 'dashboard/confirm_delete.html', {
        'object_name': workshop.title,
        'cancel_url_name': 'dashboard_workshop_list',
    })


# -------- EXPORT TO EXCEL --------

@login_required
def material_export_excel_view(request):
    if not staff_only(request):
        return redirect('profile')

    materials = Material.objects.all().order_by('name')

    # Create workbook/sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materials"

    # Header row
    headers = ["Name", "Stock Quantity", "Unit", "Description"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

    ws.freeze_panes = "A2"

    # Data rows
    row_idx = 2
    for m in materials:
        ws.cell(row=row_idx, column=1, value=m.name)
        ws.cell(row=row_idx, column=2, value=m.stock_quantity)
        ws.cell(row=row_idx, column=3, value=m.unit or "")
        ws.cell(row=row_idx, column=4, value=m.description or "")
        row_idx += 1

    # Build response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=materials.xlsx'

    wb.save(response)
    return response



#-------- Stats View --------
@login_required
def workshop_stats_view(request):
    if not staff_only(request):
        return redirect('profile')

    now = timezone.now()

    # base queryset
    qs = (
        Workshop.objects
        .select_related('place')
        .prefetch_related('materials')
        .all()
    )

    total_workshops = qs.count()
    active_workshops = qs.filter(is_active=True).count()
    inactive_workshops = qs.filter(is_active=False).count()

    upcoming_workshops = qs.filter(start_time__gte=now).count()
    past_workshops = qs.filter(end_time__lt=now).count()

    # avoid None in template: round avg capacity
    avg_capacity_val = qs.aggregate(avg_cap=Avg('capacity'))['avg_cap']
    if avg_capacity_val is None:
        avg_capacity = 0
    else:
        # round to nearest int
        avg_capacity = round(avg_capacity_val)

    # % active
    active_pct = 0
    if total_workshops > 0:
        active_pct = round((active_workshops / total_workshops) * 100)

    # workshops per place
    workshops_by_place = (
        qs.values('place__id', 'place__name')
          .annotate(workshop_count=Count('id'))
          .order_by('-workshop_count', 'place__name')
    )
    # shape: [{'place__id': 3, 'place__name': 'Library', 'workshop_count': 5}, ...]

    # workshops per month (by start_time)
    raw_month_counts = (
        qs.annotate(month=TruncMonth('start_time'))
          .values('month')
          .annotate(workshop_count=Count('id'))
          .order_by('month')
    )

    workshops_by_month = []
    for row in raw_month_counts:
        month_start = row['month'].date().replace(day=1)
        last_day_num = calendar.monthrange(month_start.year, month_start.month)[1]
        month_end = month_start.replace(day=last_day_num)

        workshops_by_month.append({
            'month': row['month'],
            'workshop_count': row['workshop_count'],
            'range_start': month_start.strftime("%Y-%m-%d"),
            'range_end': month_end.strftime("%Y-%m-%d"),
        })

    # top materials usage
    # Count how many workshops each material is attached to
    materials_usage = (
        Material.objects
        .annotate(usage_count=Count('workshops'))
        .filter(usage_count__gt=0)
        .order_by('-usage_count', 'name')
    )
    # We'll show just the top ~10 in template, but we'll pass full list
    # so you can decide in the template.

    context = {
        'now': now,

        'total_workshops': total_workshops,
        'active_workshops': active_workshops,
        'inactive_workshops': inactive_workshops,
        'upcoming_workshops': upcoming_workshops,
        'past_workshops': past_workshops,
        'avg_capacity': avg_capacity,
        'active_pct': active_pct,

        'workshops_by_place': workshops_by_place,
        'workshops_by_month': workshops_by_month,
        'materials_usage': materials_usage,
    }

    return render(request, 'dashboard/workshops/workshop_stats.html', context)