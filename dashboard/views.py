from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from accounts.models import User
from django.contrib import messages
from django.db.models import Q
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count
from datetime import timedelta


# Only staff/superusers can access this

@login_required
def dashboard_view(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return redirect('profile')
    return render(request, 'dashboard/index.html')


@login_required
def user_list_view(request):
    # only staff/superuser allowed
    if not (request.user.is_staff or request.user.is_superuser):
        return redirect('')

    # optional: search by username or email with ?q=something
    q = request.GET.get("q", "")
    if q:
        users = User.objects.filter(
            Q(username__icontains=q) | Q(email__icontains=q)
        ).order_by('id')
    else:
        users = User.objects.all().order_by('id')

    context = {
        "users": users,
        "search_query": q,
    }
    return render(request, 'dashboard/users/users.html', context)


@login_required
def user_ban_toggle_view(request, user_id):
    # only staff/superuser allowed
    if not (request.user.is_staff or request.user.is_superuser):
        return redirect('profile')

    # Only allow POST (safety: no GET side effects)
    if request.method != "POST":
        return redirect('dashboard_user_list')

    target = get_object_or_404(User, id=user_id)

    # Safety rule: don't ban yourself
    if target.id == request.user.id:
        messages.error(request, "You cannot ban yourself.")
        return redirect('dashboard_user_list')

    # Flip ban status
    target.is_banned = not target.is_banned
    target.save()

    if target.is_banned:
        messages.warning(request, f"{target.username} is now banned.")
    else:
        messages.success(request, f"{target.username} is now unbanned.")

    return redirect('dashboard_user_list')


@login_required
def user_delete_view(request, user_id):
    # only staff/superuser allowed
    if not (request.user.is_staff or request.user.is_superuser):
        return redirect('profile')

    # Only allow POST deletes
    if request.method != "POST":
        return redirect('dashboard_user_list')

    target = get_object_or_404(User, id=user_id)

    # Safety rule: don't delete yourself
    if target.id == request.user.id:
        messages.error(request, "You cannot delete your own account.")
        return redirect('dashboard_user_list')

    username = target.username
    target.delete()
    messages.error(request, f"User '{username}' has been deleted.")

    return redirect('dashboard_user_list')




# ----- EXPORT TO EXCEL -----
@login_required
def user_export_excel_view(request):
    # only staff/superuser allowed
    if not (request.user.is_staff or request.user.is_superuser):
        return redirect('profile')

    # we want export to respect the same ?q=... filter as the list view
    q = request.GET.get("q", "").strip()

    if q:
        qs = User.objects.filter(
            Q(username__icontains=q) | Q(email__icontains=q)
        ).order_by('id')
    else:
        qs = User.objects.all().order_by('id')

    # build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    headers = [
        "ID",
        "Username",
        "Email",
        "Is Staff",
        "Is Superuser",
        "Is Banned",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 24
    ws.freeze_panes = "A2"

    row_idx = 2
    for u in qs:
        ws.cell(row=row_idx, column=1, value=u.id)
        ws.cell(row=row_idx, column=2, value=u.username)
        ws.cell(row=row_idx, column=3, value=u.email)
        ws.cell(row=row_idx, column=4, value="YES" if (u.is_staff or u.is_superuser) else "no")
        ws.cell(row=row_idx, column=5, value="YES" if u.is_superuser else "no")
        ws.cell(row=row_idx, column=6, value="BANNED" if getattr(u, "is_banned", False) else "Active")
        row_idx += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=users.xlsx'

    wb.save(response)
    return response


# ----- USER STATS -----

@login_required
def user_stats_view(request):
    # only staff/superuser allowed
    if not (request.user.is_staff or request.user.is_superuser):
        return redirect('profile')

    now = timezone.now()

    qs = User.objects.all()

    total_users = qs.count()
    staff_users = qs.filter(is_staff=True).count()
    super_users = qs.filter(is_superuser=True).count()
    banned_users = qs.filter(is_banned=True).count()
    active_users = qs.filter(is_banned=False).count()  # "active" in your sense, not Django is_active

    # % banned
    banned_pct = 0
    if total_users > 0:
        banned_pct = round((banned_users / total_users) * 100)

    # new signups in last 30 days
    last_30_days = now - timedelta(days=30)
    new_last_30_days = qs.filter(date_joined__gte=last_30_days).count()

    # recently joined (top 10 newest accounts)
    # NOTE: a brand-new AbstractUser has date_joined by default
    recent_users = (
        qs.order_by('-date_joined')
          .values('id', 'username', 'email', 'date_joined', 'is_banned', 'is_staff', 'is_superuser')[:10]
    )

    # staff breakdown
    staff_breakdown = (
        qs.values('is_staff', 'is_superuser')
          .annotate(user_count=Count('id'))
          .order_by('-user_count')
    )
    # shape: [{'is_staff': True, 'is_superuser': False, 'user_count': 4}, ...]

    context = {
        'now': now,

        'total_users': total_users,
        'staff_users': staff_users,
        'super_users': super_users,
        'banned_users': banned_users,
        'active_users': active_users,
        'banned_pct': banned_pct,
        'new_last_30_days': new_last_30_days,

        'recent_users': recent_users,
        'staff_breakdown': staff_breakdown,
    }

    return render(request, 'dashboard/users/user_stats.html', context)
